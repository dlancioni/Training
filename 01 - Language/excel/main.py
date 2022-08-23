# coding=UTF-8
import os
os.system("cls || clear")

import xlrd
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook

filepath = "c:\\temp\\demo.xlsx"

def create():
    workbook = openpyxl.Workbook()
    sheet1 = workbook.create_sheet("Saldo")
    sheet2 = workbook.create_sheet("Extrato")
    workbook.save(filepath)
    
def open():   
    workbook = load_workbook(filename = filepath)
    sheet1 = workbook["Saldo"]
    sheet1.cell(1, 1).value = "David"
    workbook.save(filepath)
    
def read():   
    workbook = load_workbook(filename = filepath)
    
    sheet1 = workbook["Saldo"]
    print(sheet1.cell(2, 1).value)
    
    sheet1 = workbook["Extrato"]
    print(sheet1.cell(2, 2).value)
    print(sheet1.cell(3, 2).value)
    print(sheet1.cell(4, 2).value)
    
#create()    
read()    